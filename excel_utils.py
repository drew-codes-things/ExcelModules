"""
excel_utils.py  -  Python companion to the ExcelModules VBA collection.

Provides the same six data-cleaning / formatting operations using
openpyxl and pandas, for users who have macros disabled, use LibreOffice,
or work on macOS / Linux.

Requirements:
    pip install openpyxl pandas
    Python >= 3.8  (uses Optional[] from typing; 3.10+ X | Y syntax avoided
                   intentionally so the module works on older interpreters)

Usage:
    python excel_utils.py

Accepts .xlsx or .csv input. By default writes to a NEW file
(<original>_processed.<ext>, or _highlighted.xlsx for highlight ops) so the
original is never overwritten. Pass --output PATH to choose the output name.
"""

import argparse
import csv as csvlib
import re
import sys
from pathlib import Path
from typing import Optional

try:
    import pandas as pd
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except ImportError:
    print("Missing dependencies. Run: pip install openpyxl pandas")
    sys.exit(1)



YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
RED_FILL    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

OUTPUT_OVERRIDE: Optional[str] = None


def is_csv(path: str) -> bool:
    return Path(path).suffix.lower() == ".csv"


def resolve_output(path: str, default_suffix: str, ext: str) -> str:
    """Single source of truth for output naming: honour --output, else <stem><suffix><ext>."""
    if OUTPUT_OVERRIDE:
        return OUTPUT_OVERRIDE
    return Path(path).stem + default_suffix + ext


def resolve_output_xlsx(path: str, default_suffix: str) -> str:
    """Like resolve_output but forces .xlsx (highlight ops need cell fills)."""
    if OUTPUT_OVERRIDE:
        o = OUTPUT_OVERRIDE
        if not o.lower().endswith(".xlsx"):
            o = str(Path(o).with_suffix(".xlsx"))
        return o
    return Path(path).stem + default_suffix + ".xlsx"


def load_df(path: str, sheet: Optional[str] = None) -> "tuple[pd.DataFrame, str]":
    """Load a CSV or Excel sheet into a DataFrame. Returns (df, sheet_name)."""
    if is_csv(path):
        return pd.read_csv(path), "Sheet1"
    wb = load_workbook(path)
    sheet_name = sheet or wb.sheetnames[0]
    df = pd.read_excel(path, sheet_name=sheet_name)
    return df, sheet_name


def load_ws_workbook(path: str):
    """Return an openpyxl Workbook for cell-fill operations, building one from a CSV if needed."""
    if is_csv(path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        with open(path, newline="", encoding="utf-8") as f:
            for row in csvlib.reader(f):
                ws.append(row)
        return wb
    return load_workbook(path)


def save_output(path: str, df: pd.DataFrame, sheet: str) -> str:
    """Write df to the resolved output path (CSV in -> CSV out, unless --output says otherwise)."""
    ext = ".csv" if is_csv(path) else ".xlsx"
    out = resolve_output(path, "_processed", ext)
    if out.lower().endswith(".csv"):
        df.to_csv(out, index=False)
    else:
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name=sheet, index=False)
    return out



def sort_columns(path: str, columns: list, ascending: bool = True) -> str:
    """
    Sort the DataFrame by the given column(s) alphabetically.
    Numeric columns are skipped (same behaviour as the VBA module).
    """
    df, sheet = load_df(path)

    existing_cols = [c for c in columns if c in df.columns]
    missing_cols = [c for c in columns if c not in df.columns]
    if missing_cols:
        print(f"Ignoring unknown column(s): {missing_cols}")

    text_cols = [c for c in existing_cols if df[c].dtype == object]
    if not text_cols:
        print("No text columns found in the specified list - nothing sorted.")
        return path

    df.sort_values(by=text_cols, ascending=ascending, inplace=True, ignore_index=True)
    out = save_output(path, df, sheet)
    print(f"Sorted by {text_cols}. Output: {out}")
    return out



_FOUR_DIGIT = re.compile(r"(?<![\(\d])(\d{4})(?![\)\d])")


def brackets_around_4_digits(path: str) -> str:
    """
    Wrap standalone 4-digit numbers (e.g. years) in parentheses in every
    text cell across the first sheet.
    """
    df, sheet = load_df(path)
    changed = 0
    for col in df.select_dtypes(include="object").columns:
        original = df[col].copy()
        df[col] = df[col].apply(
            lambda v: _FOUR_DIGIT.sub(r"(\1)", str(v)) if pd.notna(v) else v
        )
        changed += (df[col] != original).sum()
    out = save_output(path, df, sheet)
    print(f"{changed} cell(s) updated. Output: {out}")
    return out



def capitalise_proper(path: str) -> str:
    """Apply Title Case to every text cell (mirrors CapitalizeFirstLetter)."""
    df, sheet = load_df(path)
    changed = 0
    for col in df.select_dtypes(include="object").columns:
        original = df[col].copy()
        df[col] = df[col].apply(lambda v: str(v).title() if pd.notna(v) else v)
        changed += (df[col] != original).sum()
    out = save_output(path, df, sheet)
    print(f"{changed} cell(s) converted. Output: {out}")
    return out


def capitalise_first_only(path: str) -> str:
    """Capitalise only the first character of each text cell."""
    df, sheet = load_df(path)
    changed = 0
    for col in df.select_dtypes(include="object").columns:
        original = df[col].copy()
        df[col] = df[col].apply(
            lambda v: str(v)[0].upper() + str(v)[1:] if pd.notna(v) and len(str(v)) > 0 else v
        )
        changed += (df[col] != original).sum()
    out = save_output(path, df, sheet)
    print(f"{changed} cell(s) updated. Output: {out}")
    return out



def highlight_duplicates_multi_column(path: str, columns: list) -> str:
    """
    Highlight rows whose combined key across `columns` is duplicated.
    Writes yellow fill to matching rows. Mirrors HighlightDuplicatesMultiColumn.
    """
    from collections import Counter
    wb = load_ws_workbook(path)
    ws = wb.active

    header = {cell.value: cell.column for cell in ws[1]}
    col_indices = [header[c] for c in columns if c in header]
    if not col_indices:
        print("No valid columns found for duplicate highlighting.")
        return path

    keys = [
        "\x00".join(str(ws.cell(r, ci).value) for ci in col_indices)
        for r in range(2, ws.max_row + 1)
    ]
    counts = Counter(keys)
    highlighted = 0
    for i, key in enumerate(keys):
        if counts[key] > 1:
            for ci in col_indices:
                ws.cell(i + 2, ci).fill = YELLOW_FILL
            highlighted += 1

    out = resolve_output_xlsx(path, "_highlighted")
    wb.save(out)
    print(f"{highlighted} row(s) highlighted. Output: {out}")
    return out


def highlight_duplicates_across_sheets(path: str, col_name: str, sheet_a: str, sheet_b: str) -> str:
    """
    Highlight cells in sheet_a[col_name] that also appear in sheet_b[col_name].
    Mirrors HighlightDuplicatesAcrossSheets.
    """
    if is_csv(path):
        print("Cross-sheet highlight needs a multi-sheet .xlsx, not a CSV.")
        return path
    wb = load_workbook(path)
    ws_a = wb[sheet_a]
    ws_b = wb[sheet_b]

    header_a = {cell.value: cell.column for cell in ws_a[1]}
    header_b = {cell.value: cell.column for cell in ws_b[1]}
    ci_a = header_a.get(col_name)
    ci_b = header_b.get(col_name)
    if not ci_a or not ci_b:
        print(f"Column '{col_name}' not found in one or both sheets.")
        return path

    vals_b = {ws_b.cell(r, ci_b).value for r in range(2, ws_b.max_row + 1)}
    highlighted = 0
    for r in range(2, ws_a.max_row + 1):
        cell = ws_a.cell(r, ci_a)
        if cell.value in vals_b:
            cell.fill = RED_FILL
            highlighted += 1

    out = resolve_output_xlsx(path, "_cross_highlighted")
    wb.save(out)
    print(f"{highlighted} cell(s) highlighted in red. Output: {out}")
    return out



def remove_exact_duplicates(path: str, columns: Optional[list] = None,
                            keep: str = "first") -> str:
    """
    Remove duplicate rows. `keep` can be 'first', 'last', or False (remove all).
    Mirrors RemoveExactDuplicates / RemoveDuplicatesKeepFirst.
    """
    df, sheet = load_df(path)
    before = len(df)
    df.drop_duplicates(subset=columns, keep=keep, inplace=True, ignore_index=True)
    removed = before - len(df)
    out = save_output(path, df, sheet)
    print(f"{removed} duplicate row(s) removed. Output: {out}")
    return out



def strip_whitespace(path: str, remove_all: bool = False) -> str:
    """
    Trim leading/trailing (and double) spaces from every text cell.
    If remove_all=True, strips ALL spaces (mirrors RemoveAllSpaces).
    """
    df, sheet = load_df(path)
    changed = 0
    for col in df.select_dtypes(include="object").columns:
        original = df[col].copy()
        if remove_all:
            df[col] = df[col].apply(lambda v: str(v).replace(" ", "") if pd.notna(v) else v)
        else:
            df[col] = df[col].apply(lambda v: " ".join(str(v).split()) if pd.notna(v) else v)
        changed += (df[col] != original).sum()
    out = save_output(path, df, sheet)
    print(f"{changed} cell(s) updated. Output: {out}")
    return out



MENU = """
============================
  ExcelModules - Python CLI
============================
 1  Alphabetical sort
 2  Brackets around 4-digit numbers
 3  Capitalise (Proper Case)
 4  Capitalise (first letter only)
 5  Highlight duplicates - multi-column
 6  Highlight duplicates - across two sheets
 7  Remove duplicate rows (keep first)
 8  Remove ALL duplicate rows
 9  Strip whitespace (trim)
 10 Strip whitespace (remove all spaces)
 0  Exit
"""


def parse_args():
    p = argparse.ArgumentParser(description="ExcelModules - Python CLI (xlsx/csv data cleaning).")
    p.add_argument("-i", "--input", default=None, help="Path to the .xlsx or .csv file (skips the prompt)")
    p.add_argument("-o", "--output", default=None,
                   help="Exact output path for the result (overrides the default _processed/_highlighted naming)")
    return p.parse_args()


def main():
    global OUTPUT_OVERRIDE
    args = parse_args()
    OUTPUT_OVERRIDE = args.output

    print(MENU)
    path = args.input or input("Enter path to your .xlsx or .csv file: ").strip().strip('"')
    if not Path(path).exists():
        print("File not found.")
        return

    choice = input("Choose an option: ").strip()

    if choice == "1":
        cols = input("Column names to sort by (comma-separated): ").split(",")
        asc  = input("Ascending? (y/n): ").lower() != "n"
        sort_columns(path, [c.strip() for c in cols], ascending=asc)
    elif choice == "2":
        brackets_around_4_digits(path)
    elif choice == "3":
        capitalise_proper(path)
    elif choice == "4":
        capitalise_first_only(path)
    elif choice == "5":
        cols = input("Column names to check (comma-separated): ").split(",")
        highlight_duplicates_multi_column(path, [c.strip() for c in cols])
    elif choice == "6":
        col   = input("Column name to compare: ").strip()
        sh_a  = input("Sheet A name: ").strip()
        sh_b  = input("Sheet B name: ").strip()
        highlight_duplicates_across_sheets(path, col, sh_a, sh_b)
    elif choice == "7":
        cols = input("Columns to deduplicate on (comma-separated, or Enter for all): ").strip()
        remove_exact_duplicates(path, columns=[c.strip() for c in cols.split(",")] if cols else None, keep="first")
    elif choice == "8":
        cols = input("Columns to deduplicate on (comma-separated, or Enter for all): ").strip()
        remove_exact_duplicates(path, columns=[c.strip() for c in cols.split(",")] if cols else None, keep=False)
    elif choice == "9":
        strip_whitespace(path)
    elif choice == "10":
        strip_whitespace(path, remove_all=True)
    elif choice == "0":
        sys.exit(0)
    else:
        print("Invalid choice.")


if __name__ == "__main__":
    main()
